from tkinter import *
from tkinter import ttk
import tempfile
import os
import xlwt
import xlrd
import openpyxl
import random
import tkinter.messagebox
from datetime import datetime;

root = Tk()
root.geometry("1600x800+0+0")
root.title("Kanaka Durga Stationary and General Merchants")

text_Input = DoubleVar()    #code name for Total
operator = 0    #Total
Warn = StringVar()
Warnoperator = ""
temp1 = []      #index of item sold
temp2 = []      #quantity of item sold

path = "itemlist.xlsx"
inputWorkbook = xlrd.open_workbook(path)
inputWorksheet = inputWorkbook.sheet_by_index(0)
n_rows = inputWorksheet.nrows
n_cols = inputWorksheet.ncols

A = []  #Names of items
B = []  #Sell price
C = []  #minimum quantity
D = []  #Quantity
for i in range(1,n_rows):
    A.append(inputWorksheet.cell_value(i,1))
    B.append(inputWorksheet.cell_value(i,4))
    D.append(inputWorksheet.cell_value(i,2))
    C.append(inputWorksheet.cell_value(i,5))

##A = ["apple","mango","grape","pineapple","custurd apple","papaya"]
##B = [10,12,13,14,15,16]
##C = [1,2,3,4,5,101]   #minimum
##D = [100,100,100,100,100,100]

f = Frame(root,bg = "powder blue",bd=20,relief=RIDGE)
f.grid()
        
f1 = Frame(f,bd =14, width = 1500, height = 100, padx= 10,bg = "powder blue", relief = RIDGE)
f1.grid(row = 0,column = 0 ,columnspan =4,sticky=W)
        
f5 = Frame(f,bd =14, width = 1500, height = 550, padx= 10,bg = "cadet blue", relief = RIDGE)
f5.grid(row = 1,column = 0 ,sticky=W)
        
f4 = Frame(f,bd =14, width = 1500, height = 200, padx= 10,bg = "cadet blue", relief = RIDGE)
f4.grid(row = 2,column = 0 ,sticky=W)
        
f2 = Frame(f5,bd =14, width = 875, height = 550, padx= 10,bg = "powder blue", relief = RIDGE)
f2.grid(row = 0,column = 0 ,sticky=W)

f3 = Frame(f5,bd =14, width = 575, height = 550, padx= 10,bg = "cadet blue", relief = RIDGE)
f3.grid(row = 0,column = 1 ,sticky=W)

Date1 = StringVar()
Date1.set(datetime.now().strftime("%d/%m/%Y"))

lblTitle = Label(f1, textvariable = Date1, font = ('arial',30,'bold'),pady=9,bg = "powder blue",fg = "coral1").grid(row=0,column=0)
lblTitle = Label(f1, text = "\t\tCustomer Billing System\t\t\t\t", font = ('arial',30,'bold'),pady = 9,bd = 5,bg = "powder blue",fg = "coral1",justify = CENTER).grid(row=0,column=1)

def iadd():
    global operator
    global temp1
    global temp2
    result = "  "
    searched_label = Label(f2,text = result,bg = "powder blue")
    searched_label.grid(row=2,column =1,padx = 10)
    tempvar1 = float(Quantity.get())
    tempvar2 = str(Name.get())
    index = 0
    for j in A:
        if(j!=tempvar2):
                index = index+1
        else:
                index = index
                break
    if index == len(A):
        result = "Record not found"
        searched_label = Label(f2,text = result,bg="powder blue")
        searched_label.grid(row=2,column =1,padx = 10 )
    else:
        result = B[index]*tempvar1
        temp1.append(int(index))
        temp2.append(tempvar1)
        operator = operator + result
        text_Input.set(operator)
        txtReceipt.insert(END,str(tempvar2)+"  "+str(tempvar1)+"\t\t\t\t\t\t"+str(result)+"\n")    

def ireset():
    global operator
    global temp1
    global temp2
    temp1 = []
    temp2 = []
    operator = 0 
    text_Input.set(0)
    txtReceipt.delete('1.0',END)
    txtReceipt.insert(END,Date1.get()+"\t\t              "+"Kanaka Durga Stationary                "+"\t\t\t"+"ph.no: 924678813"+"\n\n")
    txtReceipt.insert(END, "Items\t\t\t\t\t\t"+"Cost of item\n\n")
    

def iprint():
    global Warnoperator
    global temp1
    global temp2
    txtReceipt.insert(END,"\n"+"Total"+"\t\t\t\t\t\t"+str(operator)+"\n")
    q= txtReceipt.get("1.0","end-1c")
    filename = tempfile.mktemp(".txt")
    open(filename,"w").write(q)
    os.startfile(filename,"print")
    j = 0
    for i in D:
        if D[j]<=C[j]:
            Warnoperator = Warnoperator + A[j] + " "
            Warn.set(Warnoperator)
        j = j+1
    wbkName = "itemlist.xlsx"
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        dummy = 0
        for i in temp1:
            wks.cell(i+2,3).value = wks.cell(i+2,3).value-temp2[dummy]
            dummy = dummy +1
            if wks.cell(i+2,3).value <= C[i]:
                Warnoperator = Warnoperator + A[i] + " "
                Warn.set(Warnoperator)
    wbk.save(wbkName)
    wbk.close
    
            
    

def ibalanceitems():
    global n_rows
    #global Warnoperator
    Warn.set("")
##    iexit = tkinter.messagebox.askyesno("Customer Billing System","Confirm if you want to exit")
##    if iexit > 0:
##        root.destroy()
##        return
    wbkName = "itemlist.xlsx"
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        dummy = ""
        for i in range(1,n_rows):
            if wks.cell(i+1,3).value <= wks.cell(i+1,6).value:
                dummy = dummy + str(wks.cell(i+1,2).value) + " "
##            wks.cell(i+2,3).value = D[i]-temp2[dummy]
##            dummy = dummy +1
    Warn.set(dummy)
    wbk.save(wbkName)
    wbk.close
    

def on_keyrelease(event):
    #value = event.widget.get()      #get text from entry
    value = Name.get()
    value = value.strip().lower()
    data = []
    if value=='':
        listbox_update(data)
    else:
        for item in A:
            if value in item.lower():
                data.append(item)

        listbox_update(data)    #update data in listbox

def listbox_update(data):
    #delete previous data
    listbox.delete(0,'end')
    #sorting data
    data = sorted(data, key = str.lower)

    # put new data
    for item in data:
        listbox.insert('end',item)

def on_select(event):
    entry.delete(0,'end')
    entry.insert(0,event.widget.get(event.widget.curselection()))
    data=[event.widget.get(event.widget.curselection())]
    Name.set(str(event.widget.get(event.widget.curselection())))
    listbox_update(data)
    listbox.delete(0,'end')

Name = StringVar()
Quantity = StringVar()


lblName = Label(f2,font= ('arial',12,'bold'),text = "Name:",fg = "Cornsilk",bg ="Cadet Blue")
lblName.grid(row=0,column=0,padx = 10)
entry = Entry(f2,font = ('arial', 12,'bold'),textvariable = Name, width = 30)
entry.grid(row = 0,column = 1,padx = 20, pady = 3)
entry.bind('<KeyRelease>',on_keyrelease)

listbox = Listbox(f2,bg ='powder blue',bd = 0,width = 45)
listbox.grid(row = 1,column = 1)
listbox.bind('<<ListboxSelect>>',on_select)

lblQuantity = Label(f2,font= ('arial',12,'bold'),text = "Quantity:",padx=2,fg = "Cornsilk",bg ="Cadet Blue")
lblQuantity.grid(row=0,column=2,padx = 10)
txtQuantity = Entry(f2,font = ('arial', 12,'bold'),textvariable = Quantity,width = 20)
txtQuantity.grid(row= 0, column=3,pady=3,padx=5)

lblWarn = Label(f4,font= ('arial',12,'bold'),text = "Warning:",padx=2,fg = "Cornsilk",bg ="Cadet Blue")
lblWarn.grid(row=0,column=0,padx = (10,1375))
txtWarn = Label(f4,font = ('arial', 12,'bold'),textvariable = Warn,padx = 2,fg = "Cornsilk", bg = "cadet blue")
txtWarn.grid(row= 1, column=0,pady=20,padx=1)

txtDisplay  = Entry(f2, font = ('arial', 11, 'bold'), textvariable=text_Input, bg = "powder blue")
txtDisplay.grid(row = 5,column = 3)

#=============================================
txtReceipt = Text(f3, height = 25,width=65,bd = 1,font=('arial',12))
txtReceipt.grid(row=0,column=0)

txtReceipt.insert(END,Date1.get()+"\t\t              "+"Kanaka Durga Stationary                "+"\t\t\t"+"ph.no: 924678813"+"\n\n")
txtReceipt.insert(END, "Items\t\t\t\t\t\t"+"Cost of item\n\n")

###===========================================
btnAdd = Button(f2, padx = 5,pady=7,bd=2,fg="black",font = ('arial',8,'bold'),width = 10,height =2,bg = "powder blue", text = "ADD",command = lambda:iadd()).grid(row=1,column =0)

btnReset = Button(f2, padx = 5,pady=7,bd=2,fg="black",font = ('arial',8,'bold'),width = 10,height =2,bg = "powder blue", text = "Reset", command = ireset).grid(row=1,column =2)

btnPrint = Button(f2, padx = 5,pady=7,bd=2,fg="black",font = ('arial',8,'bold'),width = 10,height =2,bg = "powder blue", text = "Print", command = iprint).grid(row=1,column =3)

btnExit = Button(f2, padx = 5,pady=7,bd=2,fg="black",font = ('arial',8,'bold'),width = 10,height =2,bg = "powder blue", text = "Notice",command = ibalanceitems).grid(row=1,column =4)

root.mainloop()


